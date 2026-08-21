# apps/reportes_app/services/excel_exporter.py
"""
Exportador Excel (.xlsx) usando openpyxl (stdlib-compatible, sin deps extras).
Si openpyxl no está instalado, devuelve CSV como fallback con extensión .xlsx.
"""

import io
import csv
import logging

logger = logging.getLogger(__name__)

# Paleta institucional
COLOR_HEADER   = '1E3A8A'   # azul oscuro
COLOR_SUBHEAD  = '3B5FC0'   # azul medio
COLOR_CRITICO  = 'B91C1C'   # rojo
COLOR_ALTO     = 'C2410C'   # naranja oscuro
COLOR_MEDIO    = 'B45309'   # amarillo
COLOR_BAJO     = '15803D'   # verde


def _nivel_color(nivel: str) -> str:
    m = {'critico': COLOR_CRITICO, 'alto': COLOR_ALTO, 'medio': COLOR_MEDIO, 'bajo': COLOR_BAJO}
    return m.get((nivel or '').lower(), '374151')


def export_excel(datos: dict) -> bytes:
    """
    Genera un archivo .xlsx a partir del dict de datos del report_generator.
    Devuelve bytes listos para HttpResponse.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        return _build_xlsx(datos, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter)
    except ImportError:
        logger.warning('openpyxl no disponible; exportando CSV como fallback')
        return _fallback_csv(datos)


def _build_xlsx(datos, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    tipo  = datos.get('tipo', '')
    filas = datos.get('filas', [])

    # ── Título principal ───────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    titulo_cell = ws['A1']
    titulo_cell.value = datos.get('titulo', 'Reporte SGEP')
    titulo_cell.font      = Font(bold=True, size=14, color='FFFFFF')
    titulo_cell.fill      = PatternFill('solid', fgColor=COLOR_HEADER)
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    gen_cell = ws['A2']
    gen_cell.value = f'Generado: {datos.get("generado", "")}  |  Sistema SGEP'
    gen_cell.font      = Font(size=10, color='FFFFFF', italic=True)
    gen_cell.fill      = PatternFill('solid', fgColor=COLOR_SUBHEAD)
    gen_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    row = 4  # Fila de inicio de datos

    # ── Resumen estadístico ────────────────────────────────────────────────────
    resumen = _build_resumen(datos)
    if resumen:
        ws.cell(row, 1, 'RESUMEN').font = Font(bold=True, size=11)
        row += 1
        for k, v in resumen.items():
            ws.cell(row, 1, k).font = Font(bold=True)
            ws.cell(row, 2, str(v))
            row += 1
        row += 1

    # ── Tabla de filas ─────────────────────────────────────────────────────────
    if filas:
        headers = list(filas[0].keys())
        header_labels = {
            'nombre_completo': 'Nombre Completo', 'cedula': 'Cédula',
            'sexo': 'Sexo', 'estado': 'Estado', 'delito': 'Delito',
            'centro': 'Centro', 'celda': 'Celda', 'ingreso': 'Ingreso',
            'titulo': 'Evaluación', 'interno': 'Interno', 'psicologo': 'Psicólogo',
            'nivel_riesgo': 'Nivel Riesgo', 'calificacion': 'Calificación',
            'fecha': 'Fecha', 'programa': 'Programa', 'tipo': 'Tipo',
            'progreso': 'Progreso', 'inicio': 'Inicio', 'fin_previsto': 'Fin Previsto',
            'evento': 'Evento', 'usuario': 'Usuario', 'ip': 'IP',
            'descripcion': 'Descripción', 'score': 'Score', 'confianza': 'Confianza',
            'modelo': 'Modelo', 'inicio_real': 'Inicio Real',
        }
        # encabezados
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header_labels.get(h, h.replace('_', ' ').title()))
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill('solid', fgColor=COLOR_HEADER)
            cell.alignment = Alignment(horizontal='center')

        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

        for f_idx, fila in enumerate(filas):
            fill_color = 'F8FAFC' if f_idx % 2 == 0 else 'FFFFFF'
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row + f_idx, col_idx, fila.get(h, ''))
                cell.fill = PatternFill('solid', fgColor=fill_color)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)

        row += len(filas) + 1

    # ── Autofit columnas ───────────────────────────────────────────────────────
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_resumen(datos: dict) -> dict:
    tipo = datos.get('tipo', '')
    r = {}
    if tipo == 'internos':
        r['Total internos activos'] = datos.get('total', 0)
    elif tipo == 'evaluaciones':
        r['Total evaluaciones'] = datos.get('total', 0)
        r['Completadas'] = datos.get('completadas', 0)
        r['Pendientes'] = datos.get('pendientes', 0)
        if datos.get('promedio_riesgo'):
            r['Promedio riesgo (%)'] = datos['promedio_riesgo']
    elif tipo == 'rehabilitacion':
        r['Total asignaciones'] = datos.get('total', 0)
        r['Completados'] = datos.get('completados', 0)
        r['En proceso'] = datos.get('en_proceso', 0)
        r['Progreso promedio (%)'] = datos.get('avg_progreso', 0)
    elif tipo == 'auditoria':
        r['Total eventos'] = datos.get('total', 0)
        r['Exitosos'] = datos.get('exitosos', 0)
        r['Fallidos'] = datos.get('fallidos', 0)
    elif tipo == 'estadistico':
        r['Internos activos'] = datos.get('internos', 0)
        r['Evaluaciones'] = datos.get('evaluaciones', 0)
        r['Rehabilitaciones'] = datos.get('rehabilitaciones', 0)
        r['Predicciones IA'] = datos.get('predicciones_ia', 0)
    elif tipo == 'ia_predictiva':
        r['Total predicciones'] = datos.get('total', 0)
        r['Score promedio'] = f"{datos.get('avg_score', 0)}%"
        r['Confianza promedio'] = f"{datos.get('avg_confianza', 0)}%"
    return r


def _fallback_csv(datos: dict) -> bytes:
    """CSV como fallback si openpyxl no está disponible."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([datos.get('titulo', 'Reporte'), f"Generado: {datos.get('generado','')}"])
    writer.writerow([])
    filas = datos.get('filas', [])
    if filas:
        writer.writerow(list(filas[0].keys()))
        for fila in filas:
            writer.writerow(list(fila.values()))
    return buf.getvalue().encode('utf-8-sig')
